import os
from collections import Counter
import openpyxl
from openpyxl.styles import Font, Alignment


def most_occuring_element(lst):
    if len(lst) == 0:
        return -1, -1
    # Count occurrences of each element in the list
    count_dict = Counter(lst)

    # Find the element with the maximum count
    most_common_element = count_dict.most_common(1)[0][0]
    most_common_count = count_dict.most_common(1)[0][1]

    return most_common_element, most_common_count


def get_folder_names(path):
    folders = []

    while True:
        path, folder = os.path.split(path)
        if folder != "":
            folders.append(folder)
        else:
            if path != "":
                folders.append(path)
            break

    folders.reverse()  # Reversing the list to get the correct order
    return folders


def write_to_excel(tables, target_path, var_columns = ['A', 'B']):
    workbook = openpyxl.Workbook()
    del workbook['Sheet']
    for i, table in enumerate(tables):
        worksheet = workbook.create_sheet(str(i) + "_" + table['class'])

        # Set the width
        # for val_col in var_columns:
        worksheet.column_dimensions['A'].width = 50
        worksheet.column_dimensions['B'].width = 30

        # Create a bold font style for column headers
        bold_font = Font(bold=True)

        # Loop through the list of lists and write data to the worksheet
        for row_index, row_data in enumerate([table['header']] + table['body'], start=1):
            for col_index, cell_value in enumerate(row_data, start=1):
                cell = worksheet.cell(row=row_index, column=col_index, value=cell_value)
                # Apply bold font to the header row (first row)
                if row_index == 1:
                    cell.font = bold_font
                    cell.alignment = Alignment(wrap_text=True,vertical='top')
                if col_index == 1:
                    # cell.style.alignment.wrap_text=True
                    cell.alignment = Alignment(wrap_text=True,vertical='top')

    workbook.save(target_path)
