import json
import os
import warnings
from datetime import date, datetime
from glob import glob

warnings.filterwarnings("ignore")

import requests
from loguru import logger
from secedgar import CompanyFilings, FilingType

from classification import *
from extraction.tables import *
import openpyxl


def get_filing_urls(cik):
    filing = CompanyFilings(cik_lookup=cik,
                            filing_type=FilingType.FILING_10K,
                            start_date=date(2018, 1, 1),
                            end_date=datetime.now(),
                            user_agent="Name (email)")

    company_filings_urls = filing.get_urls()
    return company_filings_urls[cik]


def download_file(link, folder_path):
    headers = {
       "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/93.0.4577.63 Safari/537.36"
    }
    r = requests.get(link, headers=headers)
    file_name = link.split('/')[-1]
    file_path = os.path.join(folder_path, file_name)
    with open(file_path, 'wb') as f:
        f.write(r.content)
    return r.text


def download_company_files(cik):
    filing_urls = get_filing_urls(cik)
    for filing_url in filing_urls:
        html_master = "/".join(filing_url.split('/')[:-1])
        index_url = filing_url.replace(".txt", "-index-headers.html")
        folder_path = f"data/{cik}/" + html_master.split('/')[-1]
        table_json_path = os.path.join(folder_path, "tables.json")
        if os.path.exists(table_json_path):
            continue
        
        logger.info(f"processing {filing_url}")
        os.makedirs(folder_path, exist_ok=True)
        index_content = download_file(index_url, folder_path)
        soup = BeautifulSoup(index_content, "html.parser")
        links = soup.find_all("a")
        for link in links:
            if "Financial_Report.xlsx" in link['href']:
                download_file(html_master + "/" + "Financial_Report.xlsx", folder_path)

        report_html = download_file(html_master + "/" + links[0]['href'], folder_path)
        tables = get_html_tables(report_html)
        tables = classify_tables(tables)
        tables = fix_table_classification(tables)
        tables = construct_proper_tables(tables)

        data = {
            'url': filing_url,
            'tables': tables
        }
        with open(table_json_path, 'w') as f:
            json.dump(data, f, indent=4)


def reprocess_company_files(cik):
    for table_json_path in glob(f"data/{cik}/*/tables.json"):
        reprocess_file(table_json_path)


def reprocess_file(table_json_path):
    with open(table_json_path, 'r') as f:
        data = json.load(f)
    logger.info(f"processing {table_json_path}")

    tables = data['tables']
    tables = classify_tables(tables)
    tables = fix_table_classification(tables)
    tables = construct_tables(tables)
    data['tables'] = tables    

    with open(table_json_path, 'w') as f:
        json.dump(data, f, indent=4)


######################## - Excel code - ######################
def process_finance_excel(excel_path):
    logger.info(f"processing file:{excel_path}")
    workbook = openpyxl.load_workbook(excel_path)
    tables = []
    for sheet_name in tqdm(workbook.sheetnames[:20]):
        sheet = workbook[sheet_name]
        table = []
        for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row):
            table_row = []
            # ignoring foot note strings.
            if row[0].value is not None and re.search("^(\[\d{1,}\](,|\s{1,})?){1,}$", str(row[0].value).strip()):
                continue
            for cell in row:
                cell_value = cell.value
                if type(cell) == openpyxl.cell.MergedCell:
                    cell_value = find_merged_cell_value(sheet, cell.row, cell.column)
                if cell_value is None or \
                        re.search("^(\[\d{1,}\](,|\s{1,})?){1,}$", str(cell_value).strip()): # for detecting footnote numbers
                    cell_value = ''
                table_row.append(str(cell_value))
            table.append(table_row)
        tables.append(get_final_table_obj(sheet, table))
    
    return tables


def get_final_table_obj(sheet, table):
    col_index = 1
    for range_ in sheet.merged_cells.ranges:
        if (1, 1) in range_.left:
            col_index = range_.max_row
            break
    
    columns = ["" for _ in range(len(table[0]))]
    for row in table[:col_index]:
        for i in range(1,len(row)):
            columns[i] = columns[i] + " " + row[i]
    for i in range(len(columns)):
        columns[i] = columns[i].strip()
    table_body = drop_footnote_strings(table[col_index:])
    table_body = drop_empty_rows(table[col_index:])
    table_body, columns = drop_empty_columns(table_body, columns)

    table_object = {
        "title": table[0][0],
        "class": get_table_class(table[0][0]),
        "header": columns,
        "body": table_body
    }
    return table_object


def find_merged_cell_value(sheet, row, column):
    for range_ in sheet.merged_cells.ranges:
        if (row, column) in range_.top: # merging only horizontally merged cells.
            return sheet.cell(row=range_.min_row, column=range_.min_col).value
    return ''

if __name__ == "__main__":
    # ciks = ['MMM', 'ABT', 'ABBV', 'ABMD', 'ACN', 'ATVI', 'ADBE', 'AMD', 'AAP', 'AES', 'AFL', 'A', 'APD', 'AKAM', 'ALK', 'ALB', 'ARE', 'ALXN', 'ALGN', 'ALLE', 'LNT', 'ALL', 'GOOGL', 'GOOG', 'MO', 'AMZN', 'AMCR', 'AEE', 'AAL', 'AEP', 'AXP', 'AIG', 'AMT', 'AWK', 'AMP', 'ABC', 'AME', 'AMGN', 'APH', 'ADI', 'ANSS', 'ANTM', 'AON', 'AOS', 'APA', 'AAPL', 'AMAT', 'APTV', 'ADM', 'ANET', 'AJG', 'AIZ', 'T', 'ATO', 'ADSK', 'ADP', 'AZO', 'AVB', 'AVY', 'BKR', 'BLL', 'BAC', 'BK', 'BAX', 'BDX', 'BRK.B', 'BBY', 'BIO', 'BIIB', 'BLK', 'BA', 'BKNG', 'BWA', 'BXP', 'BSX', 'BMY', 'AVGO', 'BR', 'BF.B', 'CHRW', 'COG', 'CDNS', 'CPB', 'COF', 'CAH', 'KMX', 'CCL', 'CARR', 'CTLT', 'CAT', 'CBOE', 'CBRE', 'CDW', 'CE', 'CNC', 'CNP', 'CTL', 'CERN', 'CF', 'SCHW', 'CHTR', 'CVX', 'CMG', 'CB', 'CHD', 'CI', 'CINF', 'CTAS', 'CSCO', 'C', 'CFG', 'CTXS', 'CLX', 'CME', 'CMS', 'KO', 'CTSH', 'CL', 'CMCSA', 'CMA', 'CAG', 'CXO', 'COP', 'ED', 'STZ', 'COO', 'CPRT', 'GLW', 'CTVA', 'COST', 'CCI', 'CSX', 'CMI', 'CVS', 'DHI', 'DHR', 'DRI', 'DVA', 'DE', 'DAL', 'XRAY', 'DVN', 'DXCM', 'FDX', 'FIS', 'FITB', 'FE', 'FRC', 'FISV', 'FLT', 'FLIR', 'FLS', 'FMC', 'F', 'FTNT', 'FTV', 'FBHS', 'FOXA', 'FOX', 'BEN', 'FCX', 'GPS', 'GRMN', 'IT', 'GD', 'GE', 'GIS', 'GM', 'GPC', 'GILD', 'GL', 'GPN', 'GS', 'GWW', 'HRB', 'HAL', 'HBI', 'HIG', 'HAS', 'HCA', 'PEAK', 'HSIC', 'HSY', 'HES', 'HPE', 'HLT', 'HFC', 'HOLX', 'HD', 'HON', 'HRL', 'HST', 'HWM', 'HPQ', 'HUM', 'HBAN', 'HII', 'IEX', 'IDXX', 'INFO', 'ITW', 'ILMN', 'INTC', 'ICE', 'IBM', 'IP', 'IPG', 'IFF', 'INTU', 'ISRG', 'IVZ', 'IPGP', 'IQV', 'IRM', 'JKHY', 'J', 'JBHT', 'SJM', 'JNJ', 'JCI', 'JPM', 'JNPR', 'KSU', 'K', 'KEY', 'KEYS', 'KMB', 'KIM', 'KMI', 'KLAC', 'KHC', 'KR', 'LB', 'LHX', 'LH', 'LRCX', 'LW', 'LVS', 'LEG', 'LDOS', 'LEN', 'LLY', 'LNC', 'LIN', 'LYV', 'LKQ', 'LMT', 'L', 'LOW', 'LUMN', 'LYB', 'MTB', 'MRO', 'MPC', 'MKTX', 'MAR', 'MMC', 'MLM', 'MAS', 'MA', 'MKC', 'MXIM', 'MCD', 'MCK', 'MDT', 'MRK', 'MET', 'MTD', 'MGM', 'MCHP', 'MU', 'MSFT', 'MAA', 'MHK', 'TAP', 'MDLZ', 'MNST', 'MCO', 'MS', 'MOS', 'MSI', 'MSCI', 'MYL', 'NDAQ', 'NOV', 'NTAP', 'NFLX', 'NWL', 'NEM', 'NWSA', 'NWS', 'NEE', 'NLSN', 'NKE', 'NI', 'NSC', 'NTRS', 'NOC', 'NLOK', 'NCLH', 'NRG', 'NUE', 'NVDA', 'NVR', 'ORLY', 'OXY', 'ODFL', 'OMC', 'OKE', 'ORCL', 'OTIS', 'PCAR', 'PKG', 'PH', 'PAYX', 'PAYC', 'PYPL', 'PNR', 'PBCT', 'PEP', 'PKI', 'PRGO', 'PFE', 'PM', 'PSX', 'PNW', 'PXD', 'PNC', 'POOL', 'PPG', 'PPL', 'PFG', 'PG', 'PGR', 'PLD', 'PRU', 'PEG', 'PSA', 'PHM', 'PVH', 'QRVO', 'PWR', 'QCOM', 'DGX', 'RL', 'RJF', 'RTX', 'O', 'REG', 'REGN', 'RF', 'RSG', 'RMD', 'RHI', 'ROK', 'ROL', 'ROP', 'ROST', 'RCL', 'SPGI', 'CRM', 'SBAC', 'SLB', 'STX', 'SEE', 'SRE', 'NOW', 'SHW', 'SPG', 'SWKS', 'SLG', 'SNA', 'SO', 'LUV', 'SWK', 'SBUX', 'STT', 'STE', 'SYK', 'SIVB', 'SYF', 'SNPS', 'SYY', 'TMUS', 'TROW', 'TTWO', 'TPR', 'TGT', 'TEL', 'TDY', 'TFSL', 'TER', 'TSLA', 'TXN', 'TXT', 'TMO', 'TIF', 'TJX', 'TSCO', 'TT', 'TDG', 'TRV', 'TRMB', 'TFC', 'TWTR', 'TYL', 'TSN', 'UDR', 'ULTA', 'USB', 'UAA', 'UA', 'UNP', 'UAL', 'UNH', 'UPS', 'URI', 'UHS', 'UNM', 'VLO', 'VAR', 'VTR', 'VRSN', 'VRSK', 'VZ', 'VRTX', 'VIAC', 'V', 'VNO', 'VMC', 'WRB', 'WAB', 'WMT', 'WBA', 'DIS', 'WM', 'WAT', 'WEC', 'WFC', 'WELL', 'WST', 'WDC', 'WU', 'WRK', 'WY', 'WHR', 'WMB', 'WLTW', 'WYNN', 'XEL', 'XRX', 'XLNX', 'XYL', 'YUM', 'ZBRA', 'ZBH', 'ZION', 'ZTS']
    # for cik in ciks:
    #     logger.info(f"processing company, cik:{cik}")
    #     try:
    #         download_company_files(cik)
    #     except:
    #         logger.error(f"failed to process company, cik:{cik}")
    # for cik in ciks:
    #     reprocess_company_files(cik)
    # reprocess_company_files('AAP')
    # reprocess_file(r"data\aap\000115844918000039\tables.json")
    paths = glob("data/*/*/Financial_Report.xlsx")
    # paths = [r"data\AAL\000000620123000018\Financial_Report.xlsx"]
    for file in paths:
        json_path = file.replace("Financial_Report.xlsx", "tables.json")
        with open(json_path, 'r') as f:
            data = json.load(f)
        
        tables = process_finance_excel(file)
        data['tables'] = tables

        # for table in data['tables']:
        #     if 'per common share' in table['title'].lower():
        #         table['class'] = ""
        #     table['body'] = drop_empty_rows(table['body'])

        with open(json_path, "w") as f:
            json.dump(data, f, indent=4)