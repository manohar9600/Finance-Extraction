import openpyxl
from tqdm import tqdm
from extraction.tables import *


top3_table_tags = {'us-gaap:IncomeStatementAbstract': 'income statement',
                   'us-gaap:StatementOfFinancialPositionAbstract': 'balance sheet',
                   'us-gaap:StatementOfCashFlowsAbstract': 'cash flow'}


def process_finance_excel(excel_path):
    workbook = openpyxl.load_workbook(excel_path)
    tables = []
    for sheet_name in workbook.sheetnames[:20]:
        sheet = workbook[sheet_name]
        table = []
        for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row):
            table_row = []
            # ignoring foot note strings.
            if row[0].value is not None and re.search("^(\[\d{1,}\](,|\s{1,})?){1,}$", str(row[0].value).strip()):
                continue
            for cell in row:
                cell_value = cell.value
                if type(cell) == openpyxl.cell.MergedCell:
                    cell_value = find_merged_cell_value(sheet, cell.row, cell.column)
                if cell_value is None or \
                        re.search("^(\[\d{1,}\](,|\s{1,})?){1,}$", str(cell_value).strip()): # for detecting footnote numbers
                    cell_value = ''
                table_row.append(str(cell_value))
            table.append(table_row)
        tables.append(get_final_table_obj(sheet, table))
    
    return tables


def get_final_table_obj(sheet, table):
    col_index = 1
    for range_ in sheet.merged_cells.ranges:
        if (1, 1) in range_.left:
            col_index = range_.max_row
            break
    
    columns = ["" for _ in range(len(table[0]))]
    for row in table[:col_index]:
        for i in range(len(row)):
            columns[i] = columns[i] + " " + row[i]
    for i in range(len(columns)):
        columns[i] = columns[i].strip()
    table_body = drop_footnote_strings(table[col_index:])
    table_body = drop_empty_rows(table_body)
    table_body, columns = drop_empty_columns(table_body, columns)

    table_object = {
        "title": table[0][0],
        "class": "",
        "header": columns,
        "body": table_body
    }
    return table_object


def find_merged_cell_value(sheet, row, column):
    for range_ in sheet.merged_cells.ranges:
        if (row, column) in range_.top: # merging only horizontally merged cells.
            return sheet.cell(row=range_.min_row, column=range_.min_col).value
    return ''


def get_all_tags(head_list, tags):
    for obj in head_list:
        # if obj[2].get("type", "") == "Monetary":
        tags.append(obj[1]['label'])
        if len(obj[3:]) > 0:
            tags = get_all_tags(obj[3:], tags)
    return tags


def get_table_tags_reported(hierarchy_data):
    # code for fetching reported tags of each table. for now considering top3 tables.
    tags_list = []
    for table in hierarchy_data:
        if table[3][1]['label'] in top3_table_tags.keys() and \
            not 'Parenthetical' in table[1]['definition']:
            tags = []
            tags = get_all_tags(table[3][3:], tags)
        # if table[3][1]['label'] not in tags_dict:
        #     tags_dict[table[3][1]['label']] = []
        # tags_dict[table[3][1]['label']] += tags
            tags_list.append([top3_table_tags[table[3][1]['label']], 
                          table[1]['definition'].split('-')[-1].strip(), tags])
                              
    return tags_list


def classify_tables_excel(excel_tables, hierarchy_data):
    tags_list = get_table_tags_reported(hierarchy_data)
    excel_tables_dict = {}
    for table in excel_tables:
        for table_tags in tags_list:
            if table_tags[1] in table['title'] and not 'Parenthetical' in table['title']:
                if table_tags[0] not in excel_tables_dict:
                    excel_tables_dict[table_tags[0]] = []
                excel_tables_dict[table_tags[0]].append(table)
                break
    return excel_tables_dict


if __name__ == "__main__":
    from glob import glob

    paths = glob("data/apd/*/Financial_Report.xlsx")
    paths = [r"data\AXP\000000496223000006\Financial_Report.xlsx"]
    for file in paths:
        json_path = file.replace("Financial_Report.xlsx", "tables.json")
        with open(json_path, 'r') as f:
            data = json.load(f)
        logger.info(f"processing file:{file}")
        tables = process_finance_excel(file)
        data['tables'] = tables

        for table in data['tables']:
            if 'Discontinued Operations' in table['title']:
                table['class'] = ""
            table['body'] = drop_empty_rows(table['body'])

        # with open(json_path, "w") as f:
        #     json.dump(data, f, indent=4)